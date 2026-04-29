"""
[CONTEXT: HISTORIAL_TRANSACCIONES_V2.4]
Generación de archivos Excel del historial de transacciones (SC-07).

Tres hojas:
  Resumen        — fila por TransactionHistory con totales
  Detalle Items  — fila por item de cada historia
  Stock Mov.     — log de movimientos de stock_movements (filtrable)
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.stock_movement import StockMovement
from models.transaction_history import TransactionHistory


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def _style_header(ws, headers: List[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


def _to_dec(value) -> str:
    if value is None:
        return ""
    return str(value)


def _user_name(history: TransactionHistory) -> str:
    if history.user:
        return history.user.full_name or history.user.email or str(history.user.id)
    return ""


def build_historial_workbook(
    histories: Iterable[TransactionHistory],
    stock_movements: Iterable[StockMovement] | None = None,
) -> BytesIO:
    """
    Construye un Workbook openpyxl con 3 hojas y lo devuelve como BytesIO listo
    para retornar en una StreamingResponse.
    """
    wb = Workbook()

    # ----------- Hoja 1: Resumen -----------
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    headers = [
        "ID", "Tipo", "Numero", "Cliente", "Email",
        "Estado", "Total", "Tipo Doc", "Items",
        "Fecha Archivado", "Reactivado",
    ]
    _style_header(ws_resumen, headers)

    rows_buffer: List[TransactionHistory] = list(histories)
    for h in rows_buffer:
        snap = h.snapshot_data or {}
        ws_resumen.append([
            h.id,
            "Factura" if h.source_type == "invoice" else "Cotización",
            h.numero_documento,
            _user_name(h),
            (h.user.email if h.user else ""),
            h.status_at_archive,
            _to_dec(h.total),
            snap.get("tipo_documento") or "",
            len(h.items or []),
            h.archived_at.strftime("%Y-%m-%d %H:%M") if h.archived_at else "",
            h.reactivated_at.strftime("%Y-%m-%d %H:%M") if h.reactivated_at else "",
        ])

    # ----------- Hoja 2: Detalle Items -----------
    ws_items = wb.create_sheet("Detalle Items")
    item_headers = ["Historia ID", "Numero", "Descripcion", "Cantidad", "Precio Unitario", "Subtotal"]
    _style_header(ws_items, item_headers)
    for h in rows_buffer:
        for it in (h.items or []):
            ws_items.append([
                h.id,
                h.numero_documento,
                it.descripcion,
                it.cantidad,
                _to_dec(it.precio_unitario),
                _to_dec(it.subtotal),
            ])

    # ----------- Hoja 3: Movimientos Stock -----------
    ws_stock = wb.create_sheet("Movimientos Stock")
    stock_headers = ["ID", "Item ID", "Tipo", "Cantidad", "Ref Tipo", "Ref ID", "Notas", "Fecha"]
    _style_header(ws_stock, stock_headers)
    if stock_movements:
        for mv in stock_movements:
            ws_stock.append([
                mv.id,
                mv.catalog_item_id,
                mv.movement_type,
                mv.quantity,
                mv.reference_type or "",
                mv.reference_id or "",
                mv.notes or "",
                mv.created_at.strftime("%Y-%m-%d %H:%M:%S") if mv.created_at else "",
            ])

    # ----------- Footer/meta -----------
    meta_sheet = wb.create_sheet("Meta")
    meta_sheet.append(["Generado", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    meta_sheet.append(["Total registros", len(rows_buffer)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def filename_for_export(prefix: str = "historial") -> str:
    return f"{prefix}_{datetime.utcnow().strftime('%Y-%m-%d_%H%M')}.xlsx"
